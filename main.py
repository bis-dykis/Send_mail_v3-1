from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QTableWidgetItem
from openpyxl import load_workbook
from mail_control import send_mail, check_email
from main_w import Ui_mainWindow
from table_data import load_data
from dataframe import table_data
import pandas as pd
from PySide6.QtGui import QKeySequence
import datetime


class MainWindow(QMainWindow, Ui_mainWindow):
    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.list_btn.clicked.connect(self.fileopen)
        self.attach_btn.clicked.connect(self.openurl)
        self.send_btn.clicked.connect(self.sendmail)
        self.delete_btn.clicked.connect(self.delete_items)
        self.add_btn.clicked.connect(self.add_items)
        self.list_edit.setReadOnly(True)
        self.attach_edit.setReadOnly(True)
        self.list_table.setColumnWidth(1, 200)
        return

    def openurl(self):
        attach_url = QFileDialog.getExistingDirectory(self, 'Select Directory')
        self.attach_edit.setText(attach_url)
        return

    def fileopen(self):
        file_name = QFileDialog.getOpenFileName(self, 'Select File', '', 'Excel Files (*.xlsx)')
        self.list_edit.setText(file_name[0])
        df = pd.read_excel(file_name[0])
        load_data(df, self)
        return

    def sendmail(self):
        mail_id = self.id_edit.text()
        mail_pass = self.pass_edit.text()
        attach_url = self.attach_edit.text()
        list_name = self.list_edit.text()
        subject = self.subject_edit.text()
        body = self.body_edit.toPlainText()

        attach_url = str(attach_url)
        if mail_id == "":
            QMessageBox.information(self, 'Notice', '이메일을 입력하세요.')

        elif mail_pass == "":
            QMessageBox.information(self, 'Notice', '패스워드를 입력하세요.')

        elif attach_url == "":
            QMessageBox.information(self, 'Notice', '첨부파일 경로를 선택하세요.')

        elif list_name == "":
            QMessageBox.information(self, 'Notice', '메일 리스트를 선택하세요.')

        else:
            xlsx = load_workbook(list_name, read_only=True)
            if '리스트' not in xlsx.sheetnames:
                QMessageBox.information(self, 'Notice', '엑셀 파일에 리스트 시트가 없습니다.')
            else:
                if check_email(mail_id, mail_pass, self):
                    self.send_btn.setText("전송중")
                    self.send_btn.setDisabled(True)

                    table = self.list_table

                    # selected_row = table.currentRow()

                    for row in range(table.rowCount()):
                        name = table.item(row, 0).text()
                        email = table.item(row, 1).text()
                        table.setItem(row, 2, QTableWidgetItem("전송중"))
                        send_status = send_mail(mail_id, mail_pass, name, email, attach_url, subject, body)
                        table.setItem(row, 2, QTableWidgetItem(send_status))

                        if name is None:
                            break

                    self.send_btn.setText("완료")
                    df2 = pd.DataFrame()
                    df2 = table_data(df2, self)
                    print(df2)
                    now = datetime.datetime.now()
                    filename = now.strftime("%Y-%m-%d_%H-%M-%S")
                    df2.to_json(f'./result-{filename}.json', orient='records', force_ascii=False)
                    QMessageBox.information(self, 'Notice', '전송 완료')
                    self.send_btn.setDisabled(False)
                    self.send_btn.setText("메일 전송")

                    # 계속 해서 보낼 수 있도록 주석 처리
                    # self.list_edit.setText("")
                    # self.attach_edit.setText("")

                else:
                    #QMessageBox.information(self, 'Notice', '로그인 에러')
                    print('로그인 에러')
                    self.pass_edit.setFocus()

        return

    def delete_items(self):
        table = self.list_table
        selected_rows = []

        # 선택된 행의 인덱스를 담은 리스트를 생성합니다.
        for index in table.selectedIndexes():
            if index.row() not in selected_rows:
                selected_rows.append(index.row())

        # 선택된 행을 역순으로 정렬합니다.
        selected_rows.sort(reverse=True)

        # 선택된 행을 모두 삭제합니다.
        for row in selected_rows:
            table.removeRow(row)

        return

    def add_items(self):
        table = self.list_table
        rowposition = table.rowCount()
        table.insertRow(rowposition)

        return

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Copy):
            self.copy()
        elif event.matches(QKeySequence.Paste):
            self.paste()
        else:
            super().keyPressEvent(event)

    def copy(self):
        # 현재 선택한 셀의 값을 가져와서 클립보드에 복사합니다.
        selected = self.list_table.selectedRanges()[0]
        text = ''
        for i in range(selected.topRow(), selected.bottomRow() + 1):
            for j in range(selected.leftColumn(), selected.rightColumn() + 1):
                text += str(self.list_table.item(i, j).text()) + '\t'
            text = text[:-1] + '\n'
        QApplication.clipboard().setText(text)

    def paste(self):
        # 클립보드에서 텍스트를 가져옵니다.
        clipboard = QApplication.clipboard()
        text = clipboard.text().rstrip('\n')
        print(text)
        # 텍스트를 탭으로 분리합니다.
        rows = text.split('\n')

        for i, row in enumerate(rows):
            cols = row.split('\t')

            # 행 추가
            self.list_table.insertRow(self.list_table.rowCount())

            # 열 추가
            for j, col in enumerate(cols):
                item = QTableWidgetItem(col)
                self.list_table.setItem(i, j, item)
                self.list_table.setItem(self.list_table.rowCount() - 1, j, QTableWidgetItem(col))

        return

if __name__ == "__main__":
    app = QApplication()
    window = MainWindow()
    window.show()
    app.exec()
